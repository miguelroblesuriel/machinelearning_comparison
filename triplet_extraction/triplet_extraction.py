import json
import os
import time
import subprocess
import pandas as pd


from pyopenms import *

from massql import msql_fileloading

from comparison.find_triplet import find_triplet

from preprocessing.check_file_quality import check_file_quality
from preprocessing.determine_time_unit import determine_time_unit

from openpyxl import load_workbook

from comparison.compare_two_scans import compare_two_scans

def _determine_scan_polarity_pyteomics_mzML(spec):
    """
    Gets an enum for positive and negative polarity, for pyteomics

    Args:
        spec ([type]): [description]

    Returns:
        [type]: [description]
    """
    polarity = 0

    if "negative scan" in spec:
        polarity = 2
    if "positive scan" in spec:
        polarity = 1

    return polarity


def feature_finding(input_filename,input_filepath):
    """options = PeakFileOptions()
    options.setMSLevels([1])
    fh = MzMLFile()
    fh.setOptions(options)

    # Load data
    input_map = MSExperiment()
    fh.load(input_filepath, input_map)
    input_map.updateRanges()

    mtd = MassTraceDetection()
    input_mtraces: List[Kernel_MassTrace] = []

    params = mtd.getDefaults()  # Devuelve un Param con valores por defecto
    params.setValue("max_trace_length", 500)  # máximo tamaño de trace
    params.setValue("min_trace_length", 3)  # mínimo de puntos
    params.setValue("min_intensity", 1000.0)  # umbral de intensidad

    # Aquí necesitamos usar la versión que acepta Param
    # NOTA: Dependiendo de la versión de PyOpenMS, la función puede llamarse 'runExperiment'
    mtd.runExperiment(input_map, input_mtraces, params)

    # Nuevo algoritmo para metabolómica
    features = FeatureMap()

    ffm = FeatureFindingMetabo()
    params = ffm.getDefaults()
    ffm.setParameters(params)

    output_chromatograms: List[List[MSChromatogram]] = []

    ffm.run(input_mtraces, features, output_chromatograms)

    features.setUniqueIds()

    output_filename = input_filename.replace(".mzML", "_output.featureXML")
    output_path = os.path.join(
        "/mnt/d/triplet_data2/",
        output_filename
    )

    FeatureXMLFile().store(output_path, features)

    return features"""
    options = PeakFileOptions()
    options.setMSLevels([1])
    fh = MzMLFile()
    fh.setOptions(options)

    # Load data
    input_map = MSExperiment()
    fh.load(input_filepath, input_map)
    input_map.updateRanges()

    # Crear FeatureFinder
    ff = FeatureFinderAlgorithmPicked()
    fm = FeatureMap()
    seeds = FeatureMap()  # vacío si no tienes semillas
    params = Param()

    # Ejecutar detección
    ff.run(input_map, fm, params, seeds)

    fm.setUniqueIds()

    output_filename = input_filename.replace(".mzML", "_output.featureXML")
    output_path = os.path.join(
        "/mnt/d/triplet_data2/",
        output_filename
    )

    FeatureXMLFile().store(output_path, fm)
    return fm


def add_to_excel(excel_file, data_dict, filename):

    if not os.path.exists(excel_file):
        # Create a new Excel file if it doesn't exist
        df = pd.DataFrame(columns=["Metric"])
        df.to_excel(excel_file, index=False, sheet_name="Timing Data")
    # Load the existing workbook
    workbook = load_workbook(excel_file)
    sheet = workbook["Timing Data"]

    # Add the data for each metric
    for metric, value in data_dict.items():
        # Find the row for this metric
        if sheet.cell(row=1, column=1).value == "Metric":
            # Metrics are already written in the first column, so find the row for this metric
            metric_row = None
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value == metric:
                    metric_row = row
                    break
            if not metric_row:
                # If the metric is not found, add it
                metric_row = sheet.max_row + 1
                sheet.cell(row=metric_row, column=1, value=metric)

        # Find the next available column for this file and add the data
        next_col = sheet.max_column + 1
        sheet.cell(row=metric_row, column=next_col, value=value)
    sheet.cell(row=1, column=next_col, value=filename)
    # Save the workbook
    workbook.save(excel_file)


def adjust_time_unit(time_unit,rt):
    if(time_unit=="second"):
        return rt/60
    else:
        return rt

def triplet_extraction(input_filename,input_filepath,threshold,peak_threshold,find_triplet_bool=True,check_quality=True):
    start_time = time.time()
    excel_file = 'C:/Users/usuario/Desktop/Uni/import/machinelearning_comparison/triplet_data/Extra_data.xlsx'
    time_unit=determine_time_unit(input_filepath)
    print(time_unit)
    ms1_df, ms2_df = msql_fileloading.load_data(input_filepath, cache='feather')
    load_time = time.time()
    featurexml_filename = input_filename.replace(".mzML", "_output.featureXML")

    current_directory = os.getcwd()

    featurexml_file = os.path.join(current_directory, "triplet_data2", featurexml_filename)
    print(featurexml_file)
    if os.path.exists(featurexml_file):
        features = FeatureMap()
        FeatureXMLFile().load(featurexml_file, features)
    else:
        features = feature_finding(input_filename, input_filepath)

    find_feature_time = time.time()
    spectra = []
    i = 0
    duplas = []
    maxMs2i = []
    features_scans = []
    features_scans_precmz = []
    print(input_filepath)
    for f in features:
        hull = f.getConvexHull()
        (min_corner, max_corner) = hull.getBoundingBox2D()
        rt_start = min_corner[0]
        rt_end = max_corner[0]
        mz_start = min_corner[1]
        mz_end = max_corner[1]
        matched_scans = (ms2_df.loc[(
                                            (mz_start < ms2_df['precmz']) & (ms2_df['precmz'] < mz_end)) & (
                                            (rt_start / 60  < adjust_time_unit(time_unit,ms2_df['rt'])) & (adjust_time_unit(time_unit,ms2_df['rt']) < rt_end / 60 ))][
                             'scan'].unique())
        if(len(matched_scans)>0):
            features_scans.append(matched_scans)

        for m in matched_scans:
            maxMs2i.append(ms2_df.loc[ms2_df['scan'] == m]['i'].max())
        Scan_intensity_dict = {
            'scan': matched_scans,
            'maxMs2i': maxMs2i
        }
        Scan_intensity_df = pd.DataFrame(Scan_intensity_dict)
        Scan_intensity_df_sorted = Scan_intensity_df.sort_values(by='maxMs2i')
        scans = Scan_intensity_df_sorted['scan']
        """coger solo el máximo y luego uno adyacente"""
        if len(scans) > 1:
            """Sanity check: comprobar max ms2i esta en el centro y no en los laterales(no esta al borde del mz/rt)
            matched_ms1_scans = (ms1_df.loc[(
                                                (mz_start < ms1_df['mz']) & (ms1_df['mz'] < mz_end)) & (
                                                (rt_start / 60 < adjust_time_unit(time_unit, ms1_df['rt'])) & (
                                                    adjust_time_unit(time_unit, ms1_df['rt']) < rt_end / 60))][
                                 'scan'].unique())
            maxMs2i_ms1 = []
            for m in matched_ms1_scans:
                maxMs2i_ms1.append(ms2_df.loc[ms2_df['scan'] == m]['i'].max())
            Scan_ms1_intensity_dict = {
                'scan': matched_ms1_scans,
                'maxMs2i': maxMs2i_ms1
            }
            Scan_ms1_intensity_df = pd.DataFrame(Scan_ms1_intensity_dict)
            Scan_ms1_intensity_df_sorted = Scan_ms1_intensity_df.sort_values(by='maxMs2i')
            scans_ms1 = Scan_ms1_intensity_df_sorted['scan']
            print("MS1: " + str(
                ms1_df.loc[(ms1_df['i'] == ms1_df.loc[ms1_df['scan'] == scans_ms1.iloc[-1]]['i'].max()) & (
                            ms1_df['scan'] == scans_ms1.iloc[-1])][
                    'rt']) + "MS2:" + str(
                ms2_df.loc[(ms2_df['i'] == ms2_df.loc[ms2_df['scan'] == scans.iloc[-1]]['i'].max()) & (
                            ms2_df['scan'] == scans.iloc[-1])][
                    'rt']))

            print("Rt: " + str(
                ms2_df.loc[(ms2_df['i'] == ms2_df.loc[ms2_df['scan'] == scans.iloc[-1]]['i'].max()) & (
                            ms2_df['scan'] == scans.iloc[-1])][
                    'rt']) + "rt_min: " + str(rt_start) + " rt_max: " + str(rt_end))
            print("Mz: " + str(
                ms2_df.loc[(ms2_df['i'] == ms2_df.loc[ms2_df['scan'] == scans.iloc[-1]]['i'].max()) & (
                            ms2_df['scan'] == scans.iloc[-1])][
                    'precmz']) + "mz_min: " + str(mz_start) + " mz_max: " + str(mz_end))
            print("--------------------------------------------")
            

            Fin del check"""
            duplas.append(scans[-2:])
            features_scans_precmz.append(ms2_df[ms2_df['scan'] == scans.iloc[-1]]['precmz'].unique())

        spectra.clear()
        maxMs2i.clear()
        i = i + 1
    find_tuple = time.time()
    if check_quality:
        if len(duplas) > 0:
            scans = []
            for dupla in duplas:
                scans.append(dupla.iloc[0])
            check_file_quality(scans,ms2_df, input_filename)
    check_quality_time = time.time()
    triplets = []
    if find_triplet_bool:
        i = len(duplas)
        for dupla in duplas:
            print(i)
            triplet, comparison_scores = find_triplet(dupla, features_scans, features_scans_precmz, ms2_df,threshold,peak_threshold)
            print(triplet)
            triplets_dict = {
                'dupla': dupla,
                'triplet': triplet,
                'scores' : comparison_scores
            }
            triplets.append(triplets_dict)
            i = i - 1
    find_triplet_time = time.time()
    data_dict = {
        'load_file_time': start_time-load_time,
        'find_feature_time': load_time-find_feature_time,
        'find_tuple_time': find_feature_time-find_tuple,
        'check_quality_time': find_tuple-check_quality_time,
        'find_triplet_time': check_quality_time-find_triplet_time,
        'time_unit': time_unit
    }
    """add_to_excel(excel_file,data_dict,input_filename)"""
    return triplets


if __name__ == "__main__":
    print(triplet_extraction("00048166TP84_2Philic_P3b11_01_14266.mzML", "/mnt/d/filedownloads_flat/00000000poole_4_01_14268.mzML", 0.7, 5, True, False))
